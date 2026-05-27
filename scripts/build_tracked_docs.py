# Build MOVARO tracked-shop backup docs (xlsx + pdf) from the TSV, save to Desktop.
# Run locally on the Mac. Reads /tmp/tracked_shops_export.tsv (niche,name,domain,shop_id,sh_link).
import csv, os, re

TSV = "/tmp/tracked_shops_export.tsv"
DESK = "/Users/marinapetuk/Desktop"
STAMP = "2026-05-27"
BASE = "MOVARO_tracked_shops_SH10_%s" % STAMP

rows = []
with open(TSV, encoding="utf-8") as f:
    r = csv.DictReader(f, delimiter="\t")
    for x in r:
        rows.append(x)

# group preserving order
order, groups = [], {}
for x in rows:
    n = x["niche"]
    if n not in groups:
        groups[n] = []; order.append(n)
    groups[n].append(x)

# ---------- XLSX ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Tracked Shops"
hdr = ["#", "Niche", "Shop name", "Domain", "Store URL", "ShopHunter URL"]
ws.append(hdr)
hfill = PatternFill("solid", fgColor="1F2937")
for c in range(1, len(hdr) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hfill
    cell.alignment = Alignment(vertical="center")
link_font = Font(color="2563EB", underline="single")
i = 0
for niche in order:
    for x in groups[niche]:
        i += 1
        dom = (x["domain"] or "").strip()
        store_url = ("https://" + dom) if dom else ""
        ws.append([i, niche, x["name"], dom, store_url, x["sh_link"]])
        rr = ws.max_row
        if store_url:
            c = ws.cell(row=rr, column=5); c.hyperlink = store_url; c.value = dom; c.font = link_font
        c = ws.cell(row=rr, column=6); c.hyperlink = x["sh_link"]; c.value = "open in ShopHunter"; c.font = link_font
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:F%d" % ws.max_row
widths = [4, 22, 42, 30, 34, 22]
for idx, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w
# summary sheet
ws2 = wb.create_sheet("Summary")
ws2.append(["Niche", "Shops"])
for c in range(1, 3):
    ws2.cell(row=1, column=c).font = Font(bold=True)
for niche in order:
    ws2.append([niche, len(groups[niche])])
ws2.append(["TOTAL", sum(len(groups[n]) for n in order)])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)
ws2.column_dimensions["A"].width = 24; ws2.column_dimensions["B"].width = 10
xlsx_path = os.path.join(DESK, BASE + ".xlsx")
wb.save(xlsx_path)
print("XLSX:", xlsx_path)

# ---------- PDF ----------
from fpdf import FPDF

FONTS = [
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Verdana.ttf",
]
fontpath = next((p for p in FONTS if os.path.exists(p)), None)

def clean(s):
    # strip emoji / non-BMP so the PDF stays clean even if a glyph is missing
    return "".join(ch for ch in (s or "") if ord(ch) < 0x2500).strip()

pdf = FPDF(orientation="P", unit="mm", format="A4")
pdf.set_auto_page_break(auto=True, margin=15)
pdf.add_page()
if fontpath:
    pdf.add_font("U", "", fontpath)
    pdf.add_font("U", "B", fontpath)
    F = "U"
else:
    F = "Helvetica"

pdf.set_font(F, "B", 15)
pdf.cell(0, 9, clean("MOVARO — Tracked Shops Backup (ShopHunter)"), new_x="LMARGIN", new_y="NEXT")
pdf.set_font(F, "", 9)
total = sum(len(groups[n]) for n in order)
pdf.set_text_color(90, 90, 90)
pdf.cell(0, 6, clean("SH-10 · %s · %d shops · %s" % (STAMP, total,
         " | ".join("%s %d" % (n, len(groups[n])) for n in order))), new_x="LMARGIN", new_y="NEXT")
pdf.set_text_color(0, 0, 0)
pdf.ln(2)

for niche in order:
    pdf.set_font(F, "B", 12)
    pdf.set_fill_color(31, 41, 55); pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, clean("  %s — %d" % (niche, len(groups[niche]))), new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(1)
    for j, x in enumerate(groups[niche], 1):
        dom = (x["domain"] or "").strip()
        pdf.set_font(F, "B", 9)
        pdf.cell(7, 5, "%d." % j)
        pdf.set_font(F, "", 9)
        name = clean(x["name"])[:58]
        pdf.cell(86, 5, name)
        pdf.set_text_color(37, 99, 235)
        pdf.set_font(F, "", 8)
        if dom:
            pdf.cell(52, 5, dom[:34], link=("https://" + dom))
        else:
            pdf.cell(52, 5, "")
        pdf.cell(0, 5, "ShopHunter ↗", link=x["sh_link"], new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

pdf_path = os.path.join(DESK, BASE + ".pdf")
pdf.output(pdf_path)
print("PDF :", pdf_path, "| font:", fontpath or "core-Helvetica")
print("TOTAL shops:", total)
